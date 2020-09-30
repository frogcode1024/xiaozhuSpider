import requests
import time
import json
import random
import openpyxl
import xlwt
from bs4 import BeautifulSoup

def writeExcel(valueList):
    data = openpyxl.load_workbook('changsha_minsu.xlsx')
    table = data.active
    nrows = table.max_row  # 获得行数
    # ncolumns = table.max_column  # 获得列数
    # print('写入 ' + table.title + ' 第 '+ str(nrows) +' 行')
    i = 1
    for value in valueList:
        table.cell(nrows + 1, i).value = value
        i = i + 1
    data.save('changsha_minsu.xlsx')
    print("Save success!")

class XiaoZhu():
    max_num = 14

    def __init__(self):
        # self.ua = UserAgent()
        self.ua = [
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.2995.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2986.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/48.0.2564.0 Safari/537.36'
        ]
        self.temp_url = "http://cs.xiaozhu.com/yuhua-duanzufang-p{}-8/"
        self.xiaoqu = '###'

        # 获取随机的User-Agent
        self.headers = {
            "User-Agent": random.choice(self.ua)
        }
        self.parse_times = 0

    # 请求相应URL,并返回HTML文档
    def parse_url(self, url):
        response = requests.get(url, headers=self.headers)
        # 请求前睡眠2秒
        time.sleep(2)
        if response.status_code != 200:
            print('parsing not success!--', url)
            # 请求不成功
            if self.parse_times < 3:
                # 重复请求三次
                self.parse_times += 1
                return self.parse_url(url)
            else:
                # 请求不成功, parse_times置为0
                self.parse_times = 0
                return None
        else:
            # 请求成功
            print('parsing success!--', url)
            # 请求成功, parse_times重置为0
            return response.text

    # 解析列表页面，并提取详情页的URL
    def parse_html(self, html):
        soup = BeautifulSoup(html, 'lxml')
        lis = soup.select("div#page_list > ul > li")
        for li in lis:
            # 提取详情页URL
            page_url = li.select("a")[0].attrs['href']
            page_html = self.parse_url(page_url)
            item = self.parse_page(page_html)
            # self.save_item(item)
            writeExcel(item)

    # 解析详情页，并提取数据
    def parse_page(self, html):
        item_list = []
        soup = BeautifulSoup(html, 'lxml')
        temp_title = soup.select('div.pho_info > h4')[0].get_text()
        title = temp_title.replace('\n', '')
        address = soup.select("div.pho_info > p")[0].get('title')
        price = soup.select("div.day_l > span")[0].get_text()
        area_huxing = soup.select('#introduce > li.border_none > p')[0].get_text()
        people = soup.select('#introduce > li:nth-of-type(2) > h6')[0].get_text()
        bed = soup.select('#introduce > li:nth-of-type(3) > h6')[0].get_text()
        fangyuan_num = soup.select('#sameRoomNum')[0].get('value')
        host_name = soup.select("a.lorder_name")[0].get_text()
        host_gender = soup.select("div.member_pic > div")[0].get('class')[0]

        area = area_huxing.split()[0].split('：')[1]
        huxing = area_huxing.split()[1].split('：')[1] + '/' + bed + '/' + people
        name = host_name
        gender = self.gender(host_gender)

        item_list = [self.xiaoqu, title, address, area, huxing, fangyuan_num, price]
        print(item_list)
        return item_list

    # 保存数据

    def save_item(self, item_list):
        with open('XiaoZhu_chang.txt', 'a+', encoding='utf-8') as f:
            for item in item_list:
                json.dump(item, f, ensure_ascii=False, indent=2)
            f.close()
        print("Save success!")

    # 处理gender
    def gender(self, class_name):
        if class_name == 'member_ico1':
            return '女'
        if class_name == 'member_ico':
            return '男'

    # 逻辑实现
    def run(self, temp_url_xiaoqu, xiaoquName):
        # 1.Find URL
        self.xiaoqu = xiaoquName
        for i in range(1, self.max_num):
            url = temp_url_xiaoqu.format(i)
            # 2.Send Request, Get Response
            html = self.parse_url(url)
            if html:
                self.parse_html(html)
            else:
                break
        print('全部完成')

if __name__ == '__main__':
    spider = XiaoZhu()
    xiaoquList = ['yuhua', 'kaifu', 'furong', 'yuelu', 'tianxin']
    for each in xiaoquList:
        temp_url_xiaoqu = "http://cs.xiaozhu.com/" + each + "-duanzufang-p{}-8/"
        spider.run(temp_url_xiaoqu, each)

# 列表页规律
"""
http://cs.xiaozhu.com/yuhua-duanzufang-8/
http://cs.xiaozhu.com/yuhua-duanzufang-p{}-8/
"""